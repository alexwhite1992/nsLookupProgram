from dns import resolver,reversename
import openpyxl
import os


fileList = []
try:
    for filename in os.listdir('excelFiles'):
        if filename != '.DS_Store' and not filename.startswith('~$'):
            fileList.append(filename)
except:
    pass

print(fileList)

excelFile = fileList[0]
path = 'excelFiles/' + excelFile
ipList = []

myResolver = resolver.Resolver()
myResolver.nameservers = ['8.8.8.8']

wb_obj = openpyxl.load_workbook(path, data_only=True)
sheetObj = wb_obj.active

row = sheetObj.max_row

maxRow = max((c.row for c in sheetObj['A'] if c.value is not None))

print()
startingRow = int(input('What row does the data start on? '))

print('\n' + '-----------------------------------------')
print("Here are the domain names you requested: ")
print('-----------------------------------------' + '\n')

iteratingRow = startingRow
index = 1

for i in range(startingRow, maxRow + 1):
    cell_obj = sheetObj.cell(row = i, column = 1)
    ipList.append(cell_obj.value)
     
for ipaddress in ipList:
    ipaddress = ' '.join(ipaddress.split())
     
    try: 
        qname = reversename.from_address(ipaddress)
        answer = myResolver.resolve(qname, 'PTR')
        domainName = answer[0]
        domainName = str(domainName)
        domainName = domainName[:-1]
        print(str(index) + '. ' + domainName)
        sheetObj['B' + str(iteratingRow)] = domainName
        if iteratingRow <= (maxRow + 1):
            iteratingRow += 1
        index += 1

    except:        
        print(str(index) + '. ' + 'Non-existent domain')
        sheetObj['B' + str(iteratingRow)] = "Non-existent domain"
        if iteratingRow <= (maxRow + 1):
            iteratingRow += 1
        index += 1
        continue

    
    
    
wb_obj.save('excelFiles/' + excelFile)
wb_obj.save('u:/redirection/desktop/Python Programs/nslookupprogram/ns lookups/' + excelFile)
wb_obj.close()
print()
